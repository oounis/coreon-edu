#!/usr/bin/env python3
# ─────────────────────────────────────────────────────────────────────────────
# CATALOGUE PRODUIT COREON EDU — générateur
# Produit : docs/quality/Coreon-EDU_Product-Catalog.xlsx
#
# Demandé par Othman le 2026-07-25 : « une feuille Excel avec tous les éléments
# du produit, chaque élément décrit — modules, services, comptes, régions,
# demandes… — web et mobile sur la même ligne. »
#
# La feuille MODULES est construite depuis core/src/nav.js (la source unique de
# navigation) : si un module apparaît dans l'app, il apparaît ici. Les
# descriptions vivent dans ce script. Régénérer : python3 build-product-catalog.py
# ─────────────────────────────────────────────────────────────────────────────
import re, io, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, '..', '..'))

INDIGO = '4338CA'; LIGHT = 'EEF2FF'; HDR = Font(bold=True, color='FFFFFF', size=11)
FILL = PatternFill('solid', fgColor=INDIGO)

def sheet(wb, title, headers, widths, rows, note=None):
    ws = wb.create_sheet(title)
    r0 = 1
    if note:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        c = ws.cell(1, 1, note); c.font = Font(italic=True, size=10, color='555577')
        c.fill = PatternFill('solid', fgColor=LIGHT); r0 = 2
    for j, h in enumerate(headers, 1):
        c = ws.cell(r0, j, h); c.font = HDR; c.fill = FILL
    for i, row in enumerate(rows, r0 + 1):
        for j, v in enumerate(row, 1):
            c = ws.cell(i, j, v); c.alignment = Alignment(vertical='top', wrap_text=True)
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(r0 + 1, 1)
    return ws

# ── 1. MODULES depuis nav.js ────────────────────────────────────────────────
nav_src = io.open(os.path.join(ROOT, 'core/src/nav.js'), encoding='utf-8').read()
entries = re.findall(r"\{ to:'([^']+)'.*?section:'([^']+)'.*?label:'([^']+)'.*?roles:\[([^\]]*)\]", nav_src)
feat_src = io.open(os.path.join(ROOT, 'core/src/features.js'), encoding='utf-8').read()
feats_off = re.findall(r'^\s*(\w+): false', feat_src, re.M)

MOBILE = {s[:-3].lower() for s in os.listdir(os.path.join(ROOT, 'mobile/src/screens'))}
def mobile_ok(route):
    slug = route.rsplit('/', 1)[-1] or 'dashboard'
    aliases = {'app': 'dashboard', 'hr': 'staff', 'accounting': 'finance', 'academic': 'results'}
    return slug.lower() in MOBILE or aliases.get(slug, '') in MOBILE

DESC = {
 '/app': "Tableau de bord par rôle : décisions en attente, recherche (Ctrl+K), journée, Coreon Intelligence (signaux calculés des FAITS).",
 '/app/live': "Suivi en direct de la journée de l'enfant pour le parent (arrivées, repas, sieste, activités).",
 '/app/schools': "Console propriétaire : écoles clientes, coin technique par école.",
 '/app/accounts': "L'« Active Directory » de l'école : la Direction crée/gère tous les comptes ; jamais de suppression, désactivation seulement.",
 '/app/admissions': "Guichet d'admission : la candidature publique DEVIENT l'élève (docs → validation → capacité → liste d'attente), zéro ressaisie.",
 '/app/students': "Répertoire élèves + Fiche Élève 360° (identité, scolarité, finance, santé, référence ERP).",
 '/app/teachers': "Répertoire des enseignants, affectations par classe et matière.",
 '/app/staff': "Personnel non enseignant : dossiers, postes, présence.",
 '/app/hr': "RH : contrats (salaire multi-composants base/logement/transport), congés à 2 étapes, paie maker-checker, bulletin imprimable.",
 '/app/recruit': "Recrutement : reçue → entretien → offre → embauchée, refus motivé, sans saut d'étape.",
 '/app/attendance': "Présence par classe et par jour, taux, absences protégées (veille bienveillante).",
 '/app/pointage': "Pointage arrivées/départs (élèves et personnel).",
 '/app/evaluate': "Évaluation quotidienne d'une classe entière + notation par matière (grille CR-022).",
 '/app/journal': "Journal petite enfance : repas, sieste, change, humeur — ce que les ERP scolaires n'ont pas.",
 '/app/results': "Résultats et moyennes par élève, classe, matière.",
 '/app/academic': "Bulletins officiels, passage de classe, mentions par pays (CR-024), archives.",
 '/app/timetable': "Emploi du temps par classe et enseignant.",
 '/app/canteen': "Cantine : menu hebdomadaire + ALERTE ALLERGIE calculée en croisant plats et dossiers des enfants.",
 '/app/gallery': "Moments : partage photo/vidéo aux parents (l'attente n°1 d'une crèche).",
 '/app/behavior': "Comportement : encourager d'abord, tendances jamais classements (règle 9).",
 '/app/accidents': "Chaîne d'accident : schéma corporel → validation 2-yeux → accusé parent.",
 '/app/childfile': "Dossier de l'enfant : santé, vaccins, PERSONNES AUTORISÉES à récupérer, jalons.",
 '/app/security': "Sécurité : rondes, main courante, registre visiteurs.",
 '/app/requests': "Demandes : circuit approuvé → catégorie → responsable → échéance → clôture motivée + bilan du mois.",
 '/app/social': "Vie sociale de l'école : annonces, sondages, réactions.",
 '/app/events': "Événements et calendrier de l'école (fériés par pays via fetes.js).",
 '/app/accounting': "Comptabilité : barème par niveau, remises, factures, avoirs numérotés, reçus, encaissements.",
 '/app/finance': "Frais & finances familles : échéances, impayés, relances.",
 '/app/budget': "Budget & rapports : chiffres réels uniquement (recettes − paie − dépenses), dépenses annulées avec motif.",
 '/app/inventory': "Inventaire : seuils, mouvements journalisés, jamais sous zéro.",
 '/app/documents': "Documents officiels : certificats numérotés (série-année-NNNN), registre inaltérable.",
 '/app/facilities': "Location d'installations au public (piscine, stade, salles) — 2e ligne de revenus.",
 '/app/settings': "Paramètres : établissement, cycles, modules, localisation, marque, données.",
}

modules_rows = []
seen = set()
for to, section, label, roles in entries:
    if to in seen: continue
    seen.add(to)
    roles_clean = re.sub(r"'", '', roles)
    modules_rows.append([
        label, to, section, DESC.get(to, ''), roles_clean,
        'Oui', 'Oui' if mobile_ok(to) else 'Non — web seulement',
        'Actif',
    ])
for f in feats_off:
    modules_rows.append([f.capitalize(), '·', '·',
        {'homework': 'Devoirs', 'exams': 'Examens', 'library': 'Bibliothèque', 'transport': 'Transport'}.get(f, f) +
        ' — dans le code, testé, désactivé volontairement (features.js). Se rallume sans réécrire une ligne.',
        '·', 'En veille', 'En veille', 'Désactivé (choix produit)'])

# ── 2. SERVICES DU CŒUR ─────────────────────────────────────────────────────
SERVICES = [
 ('refs.js', 'Référence ERP', "Générateur unique de références (préfixe-pays-tenant-école-année-séquence-Luhn), 13 types d'entités, séquence sans trou. CR-017 : couverture partout."),
 ('acl.js + access.js', 'Contrôle d\'accès', "Refus par défaut ; blobs par rôle (le parent ne voit QUE ses enfants) ; départements cloisonnés RH / Compta / Administration (CR-019/027)."),
 ('auth.js', 'Authentification', "Connexion par rôle ; en mode serveur les mots de passe sont hachés (scrypt) et retirés de chaque blob sortant."),
 ('i18n.js', 'Bilingue FR/AR + RTL', "Le texte français EST la clé ; bascule ar/fr avec direction RTL ; test de couverture qui casse si la nav perd sa traduction."),
 ('locales.js', '4 pays (BH·QA·TN·LY)', "La couche pays : devise, pièce d'identité (CPR/QID/CIN), géographie Pays→Ville→École, cadre légal, curriculum + barème + mentions (CR-023/024). Un 5e pays = un objet de plus."),
 ('levels.js', 'Modèle de niveaux', "Capacités par niveau crèche → 6ème : ce qui s'affiche dépend de l'âge, pas d'un réglage manuel."),
 ('insights.js', 'Coreon Intelligence', "Signaux calculés des FAITS : climat de comportement, présence, recouvrement — tendances, jamais de classement d'enfants."),
 ('workbench.js', '« À décider »', "Le tableau de bord est un poste de travail : ce qui attend MA décision (admissions, accidents, congés, circuits, paiements)."),
 ('oneroster.js', 'Export OneRoster v1.2', "« Vos données sont à vous » : l'école exporte tout en un clic, au format standard ed-tech."),
 ('hr.js', 'Moteur RH', "Contrats multi-composants, congés 2 étapes, paie calculée des faits, maker-checker (préparateur ≠ validateur)."),
 ('accounting.js + budget.js', 'Moteur financier', "Barème, remises, factures/avoirs/reçus numérotés, rapport mensuel aux chiffres réels."),
 ('canteen.js', 'Moteur cantine', "Croisement menus × allergies, correspondance large (« cacahuète » → arachides)."),
 ('fetes.js', 'Calendrier des fêtes', "Fériés par pays 2026-2027 (dates religieuses « selon la lune », jamais calculées) + journées mondiales scolaires."),
 ('notify.js + mailer.js', 'Notifications', "Notifications internes par rôle + courriels d'admission."),
 ('storage.js + db.js', 'Persistance', "Écriture VÉRIFIÉE (le faux reçu est impossible) ; migrations versionnées de l'école d'exemple ; mode serveur via remote.js."),
 ('server/ (dépôt)', 'Backend v1 pilote', "Serveur Node zéro dépendance : auth scrypt, blobs par rôle, écritures verrouillées par révision, sauvegardes auto 6 h. Testé en CI, à héberger (Option 0 : Cloudflare Workers)."),
]

# ── 3. RÔLES & COMPTES ──────────────────────────────────────────────────────
ROLES = [
 ('owner', 'Propriétaire plateforme', "Console Écoles ; jamais gérable depuis une école.", 'Toutes les écoles'),
 ('schooladmin', 'Direction', "Approuve tout (paie, factures, congés, admissions) ; gère les comptes ; le DERNIER compte Direction actif est intouchable.", 'Son école'),
 ('admin', 'Administration', "Prépare et enregistre : admissions, élèves, présence, documents. NE touche NI l'argent NI la RH (CR-027).", 'Son école'),
 ('hr', 'Ressources humaines', "Possède RH, personnel, recrutement, paie (prépare ; la Direction valide).", 'Son école'),
 ('accountant', 'Comptabilité', "Possède comptabilité, frais, budget.", 'Son école'),
 ('teacher', 'Enseignant', "Évalue, fait l'appel, tient le journal, consulte ses classes.", 'Ses classes'),
 ('supervisor', 'Surveillant', "Présence, pointage, incidents.", 'Son école'),
 ('security', 'Sécurité', "Rondes, main courante, visiteurs.", 'Son école'),
 ('parent', 'Parent', "Le portail de l'enfant 0-12 : suivi direct, journal, cantine+allergies, accusés d'accident, paiements. Blob par défaut-refus.", 'SES enfants uniquement'),
]

# ── 4. PAYS & VERSIONS ──────────────────────────────────────────────────────
loc_src = io.open(os.path.join(ROOT, 'core/src/locales.js'), encoding='utf-8').read()
def country(iso):
    m = re.search(iso + r":\s*\{(.*?)\n  \w\w:", loc_src, re.S) or re.search(iso + r":\s*\{(.*)", loc_src, re.S)
    seg = m.group(1) if m else ''
    cur = re.search(r"currency:\s*'([^']+)'", seg)
    return cur.group(1) if cur else '?'
PAYS = [
 ('BH', 'Bahreïn', country('BH'), 'CPR', 'AR + FR/EN', "Marché prioritaire Golfe. Manques Tier-1 connus : passerelle Benefit, TVA 10 % NBR, Hijri, WPS/LMRA (gap-analysis)."),
 ('QA', 'Qatar', country('QA'), 'QID', 'AR + FR/EN', "Golfe. Manques Tier-1 : QPay/Fatora, e-invoicing ~2027, Hijri, WPS, reporting MOEHE."),
 ('TN', 'Tunisie', country('TN'), 'CIN', 'FR + AR', "Marché d'origine ; l'école d'exemple Al-Nour vit ici. Conformité INPDP à formaliser au premier pilote réel."),
 ('LY', 'Libye', country('LY'), 'Passeport/ID', 'AR', "Marché francophone-arabe voisin ; même couche pays, curriculum à compléter."),
]

# ── 5. INTÉGRATIONS (la Porte) ──────────────────────────────────────────────
INTEG = [
 ('Export OneRoster v1.2', 'Disponible', "Un clic, tout part au format standard. L'argument « vos données sont à vous » (l'inverse du scandale ESS/SIMS)."),
 ('Import JSON (serveur)', 'Disponible (mode serveur)', "node server/import.mjs export.json : reprise d'une école complète."),
 ('PORTE D\'INTÉGRATION SYSTÈME', 'À CONCEVOIR — CR-034', "La demande d'Othman 2026-07-25 : la porte par laquelle UN CLIENT apporte SES données (élèves, personnel, historique) — import OneRoster/CSV guidé avec correspondance des colonnes, validation, rapport d'import. Puis : API REST + webhooks (gap-analysis §4-7)."),
 ('API REST + webhooks', 'V2 (après hébergement)', "Au-delà des fichiers : intégration continue SIS/LMS/paiement/gouvernement."),
]

# ── 6. ENVIRONNEMENTS ───────────────────────────────────────────────────────
ENVS = [
 ('dev', 'Développement', "Le travail quotidien. Branche main du dépôt coreon-edu aujourd'hui ; deviendra la branche/env dev à la préparation de la version finale."),
 ('int', 'Intégration', "Les modifications demandées par les clients s'y testent AVANT prod. À créer à la préparation de la version finale."),
 ('prod', 'Production', "Ce que les clients utilisent. Livraison depuis int uniquement, jamais depuis dev. À créer à la préparation de la version finale."),
 ('—', 'Règle de flux', "dev → int → prod, jamais de saut. Une correction client se fait en int puis redescend en dev. Détail : docs/operations/environments.md"),
]

# ── Construire le classeur ──────────────────────────────────────────────────
wb = Workbook(); wb.remove(wb.active)

ws = wb.create_sheet('LISEZ-MOI')
ws.column_dimensions['A'].width = 110
lines = [
 'CATALOGUE PRODUIT — COREON EDU',
 '',
 "Tous les éléments du produit, décrits. Généré par build-product-catalog.py depuis le code",
 "(nav.js, features.js, locales.js) : si un module existe dans l'app, il est ici.",
 '',
 'Feuilles :',
 '  · Modules — chaque écran/module, ses rôles, web ET mobile sur la même ligne',
 '  · Services du cœur — les moteurs invisibles (références, ACL, i18n, paie…)',
 '  · Rôles & Comptes — qui possède quoi, qui approuve quoi',
 '  · Pays & Versions — les 4 versions : Bahreïn, Qatar, Tunisie, Libye',
 '  · Intégrations — export, import, et la Porte d\'intégration système (CR-034)',
 '  · Environnements — le plan dev / int / prod',
 '',
 'Principe web ↔ mobile : un seul cœur (core/) partagé par les deux. Une règle métier',
 "n'existe qu'une fois ; le mobile la consomme telle quelle (auto-synchro par construction).",
 '',
 'Régénérer : cd docs/quality && python3 build-product-catalog.py',
]
for i, l in enumerate(lines, 1):
    c = ws.cell(i, 1, l)
    if i == 1: c.font = Font(bold=True, size=14, color=INDIGO)

sheet(wb, 'Modules',
 ['Module', 'Route', 'Section', 'Description', 'Rôles', 'Web', 'Mobile', 'État'],
 [22, 18, 12, 58, 30, 10, 18, 22], modules_rows,
 "Web et mobile sur la MÊME LIGNE — même cœur, mêmes règles. « Non — web seulement » = écran natif restant à porter.")

sheet(wb, 'Services du cœur', ['Fichier', 'Service', 'Description'], [22, 24, 84],
 [list(s) for s in SERVICES])

sheet(wb, 'Rôles & Comptes', ['Rôle (code)', 'Nom', 'Possède / règles', 'Portée'], [14, 20, 76, 22],
 [list(r) for r in ROLES],
 "Modèle complet : docs/quality/role-model.md. Défaut-refus partout ; la Direction approuve ; maker-checker sur la paie.")

sheet(wb, 'Pays & Versions', ['ISO', 'Pays', 'Devise', 'Pièce d\'identité', 'Langues', 'Notes marché'], [8, 14, 10, 16, 14, 70],
 [list(p) for p in PAYS],
 "La couche pays (locales.js, CR-023/024) : un seul cœur, 4 configurations. Ajouter un pays = un objet de plus.")

sheet(wb, 'Intégrations', ['Canal', 'État', 'Description'], [30, 26, 76],
 [list(i) for i in INTEG])

sheet(wb, 'Environnements', ['Env', 'Rôle', 'Description'], [8, 16, 100],
 [list(e) for e in ENVS],
 "Décision Othman 2026-07-25 : 3 environnements à la préparation de la version finale ; les clients reçoivent prod.")

out = os.path.join(HERE, 'Coreon-EDU_Product-Catalog.xlsx')
wb.save(out)
print('écrit :', out, '· modules:', len(modules_rows))
