#!/usr/bin/env python3
# ════════════════════════════════════════════════════════════════════════════
# INVENTAIRE PRODUIT COREON EDU — générateur du classeur maître.
#
# Demandé par Othman le 2026-07-25 (préparation du 1er client réel, Bahreïn) :
# « un classeur Excel qui documente CHAQUE élément du système » — modules,
# pages, écrans, API, rôles, bugs, dettes, priorités, préparation production.
#
# Source des données : l'audit multi-agents du 2026-07-25 (6 audits parallèles :
# cœur A–L, cœur M–Z + serveur, pages web, mobile + parité, tests/CI/docs,
# sécurité + i18n + Bahreïn). Les constats sont figés ici À LA DATE de l'audit ;
# la colonne Statut trace ce qui a été corrigé depuis.
#
#   python3 docs/quality/build-product-inventory.py
#   → docs/quality/Coreon-EDU_Product-Inventory.xlsx
# ════════════════════════════════════════════════════════════════════════════
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), 'Coreon-EDU_Product-Inventory.xlsx')
AUDIT_DATE = '2026-07-25'
FIX_REF = 'CORRIGÉ 2026-07-25 (6dcc445)'
QA_REF = 'CORRIGÉ 2026-07-26 (FAT QA v1.4/v1.5)'

# ── style ─────────────────────────────────────────────────────────────────────
H_FILL = PatternFill('solid', fgColor='312E81')     # indigo profond
H_FONT = Font(bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(bold=True, size=14, color='312E81')
SUB_FONT = Font(size=9, italic=True, color='6B7280')
WRAP = Alignment(wrap_text=True, vertical='top')
THIN = Border(bottom=Side(style='thin', color='E5E7EB'))
SEV_FILL = {
    'CRITIQUE': PatternFill('solid', fgColor='FDE2E2'),
    'P0': PatternFill('solid', fgColor='FDE2E2'),
    'HAUTE': PatternFill('solid', fgColor='FEF3C7'),
    'P1': PatternFill('solid', fgColor='FEF3C7'),
}

def sheet(wb, name, title, subtitle, headers, widths):
    ws = wb.create_sheet(name)
    ws['A1'] = title; ws['A1'].font = TITLE_FONT
    ws['A2'] = subtitle; ws['A2'].font = SUB_FONT
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = H_FILL; c.font = H_FONT; c.alignment = WRAP
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{get_column_letter(len(headers))}4'
    return ws

def fill(ws, rows, sev_col=None):
    r = 5
    for row in rows:
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.alignment = WRAP; c.font = Font(size=9); c.border = THIN
        if sev_col:
            sev = str(row[sev_col - 1])
            for key, f in SEV_FILL.items():
                if sev.startswith(key):
                    for i in range(1, len(row) + 1):
                        ws.cell(row=r, column=i).fill = f
                    break
        r += 1

wb = Workbook(); wb.remove(wb.active)

# ═══ 1. LISEZ-MOI ═════════════════════════════════════════════════════════════
ws = wb.create_sheet('Lisez-moi')
ws.column_dimensions['A'].width = 110
lines = [
    ('INVENTAIRE PRODUIT — COREON EDU', TITLE_FONT),
    (f'Classeur maître de gestion produit & QA. Audit multi-agents du {AUDIT_DATE} '
     '(6 audits parallèles : cœur A–L, cœur M–Z + serveur, pages web, mobile + parité, '
     'tests/CI/docs, sécurité + i18n + Bahreïn).', None),
    ('', None),
    ('CONTEXTE : premier client de production = une école au BAHREÏN (gratuite 1 an). '
     'Produit livré en 4 pays : BH · QA · TN · LY. Web + mobile doivent rester synchrones.', None),
    ('', None),
    ('FEUILLES :', Font(bold=True)),
    ('  Modules (cœur)     — les 51 modules core/src : rôle, dépendances, tests, manques, préparation', None),
    ('  Pages Web          — les ~50 pages app/src/pages : rôles, défauts UX, manques école réelle', None),
    ('  Écrans Mobile      — les 28 écrans + leur état', None),
    ('  Parité Web-Mobile  — module par module : complet / partiel / MANQUANT', None),
    ('  API Serveur        — chaque endpoint : auth, rôle, limite de débit, notes', None),
    ('  Bugs & Risques     — le registre consolidé, sévérité + statut (corrigé / ouvert)', None),
    ('  Bahreïn Go-Live    — ce qui bloque ou gêne le lancement du 1er client, classé', None),
    ('  Tests & CI         — couverture, trous, faiblesses de la chaîne', None),
    ('  Dettes & Décisions — dettes documentées encore ouvertes + décisions à prendre', None),
    ('', None),
    ('LÉGENDE : Préparation 1–5 (5 = prêt production). Priorité P0 (bloqueur) → P3 (confort). '
     'Statut : OUVERT · ' + FIX_REF + ' · PLANIFIÉ (référence CR).', None),
    ('', None),
    ('RÉGÉNÉRER : python3 docs/quality/build-product-inventory.py '
     '(les données vivent dans le script ; mettre à jour le script = mettre à jour le classeur).', None),
    ('Jumeaux : Coreon-EDU_QA-Test-Workbook.xlsx (CR + tests) · Coreon-EDU_Product-Catalog.xlsx '
     '(catalogue fonctionnel, régénéré du code).', None),
]
for i, (txt, f) in enumerate(lines, 1):
    ws.cell(row=i, column=1, value=txt).alignment = WRAP
    if f: ws.cell(row=i, column=1).font = f

# ═══ 2. MODULES (CŒUR) ════════════════════════════════════════════════════════
M = [  # (module, rôle, tests, manques/bugs clés, prep 1-5, priorité, statut)
 ('academic.js','Bulletins, moyennes, passage d’année, retrait/archivage','6 tests saisie + règle n°5 ; previewPromotion/runPromotion SANS test','CAPACITY=24 dupliqué avec admissions ; 2 relectures db()/élève ; id rp collisionnable ; passage n’émet ni radiation ni annulation de factures ; TERMS = 3 trimestres fixes (BH = 2 semestres)',3,'P1','OUVERT'),
 ('access.js','Table route→rôles + liens sûrs','2 tests refus par défaut','Correspondance EXACTE seulement : /app/eleve/:id refusé → liens profonds morts ; hr/accountant privés de /app/students',3,'P1','OUVERT'),
 ('accidents.js','Chaîne accident : corps → 2 regards → parent → accusé','2 tests bureau','pendingAck NaN après rechargement (bug Date) ; SAMU 190 tunisien en dur ; approve sans contrôle de rôle ; signature parent non attribuable',3,'P0','OUVERT (bug Date = chantier now())'),
 ('accounting.js','Barème, remises, factures/avoirs, encaissements','règle n°4 avoir ; collect/dueFor/financials SANS test','PAS d’échéance / d’impayé / de plan de paiement ; numérotation = length+1 (réémission après restauration) ; 1 relecture db()/élève ; PAS de TVA ; loyers installations hors livre ; flottants pour BHD',2,'P0','OUVERT'),
 ('accounts.js','Annuaire des comptes école','3 tests','Rôles hr/accountant NON créables ; pw défaut 1234 ; reset 6 car. Math.random ; champs cin/governorate ignorent le pack',2,'P0',FIX_REF + ' (rôles) ; reste OUVERT (pw)'),
 ('acl.js','ACL serveur lecture/écriture par rôle','4 tests','READ_STRIP.admin=[] → l’Administration LIT paie/factures/salaires ; mergeWrite « * » peut EFFACER les mots de passe serveur ; blob parent fuit settings + journal/moments de la classe (fratries)',3,'P0','OUVERT'),
 ('admissions-mail.js','Emails candidat par étape (pur)','2 tests','FR uniquement ; signature de repli « Votre établissement » ; montre l’id interne au lieu de la référence',3,'P1','OUVERT'),
 ('admissions.js','Pipeline candidature → inscription','3 tests','advance() TypeError sur étape inconnue ; DOCS tunisiens non pack ; CAPACITY dupliqué ; pas de promotion de liste d’attente',3,'P1','OUVERT'),
 ('auth.js','Session 8 h par la couture stockage','1 test','Comparaison de mot de passe EN CLAIR (mode démo) ; loginAs = primitive d’usurpation exportée ; throw si user sans email ; TTL non glissant',1,'P0','OUVERT (mode démo par design — voir hébergement)'),
 ('behavior.js','Observations positives d’abord, climat de classe','4 tests','Bug Date → climat VIDE sur données réelles ; parent lié par childIds seul jamais notifié ; entriesFor exporté non filtré',3,'P1','OUVERT (bug Date = chantier now())'),
 ('budget.js','Livre de dépenses + rapports réels','1 test','Salaires sous-évalués si prime (bug setBonus — corrigé) ; 24 relectures db()/graphique ; pas de prévisionnel',3,'P2',FIX_REF + ' (via hr.js)'),
 ('canteen.js','Menu semaine + ALERTE ALLERGIE calculée','3 tests','Correspondance FRANÇAIS uniquement — « Peanut allergy » / arabe = AUCUNE alerte (sécurité enfant !) ; faux positif « Peau sensible »→gluten ; jours lun–ven (pas de dimanche BH) ; ignore childcare.healthOf',2,'P0','OUVERT'),
 ('charts.js','Jetons graphiques Recharts','AUCUN','Objets Recharts dans le cœur « neutre » ; règle 7e série non implémentée',3,'P3','OUVERT'),
 ('childcare.js','Dossier enfant : vaccins, personnes autorisées, jalons','règle n°1','VACCINS = calendrier TUNISIEN (exposition réglementaire BH) ; monthsOld ±30 j ; remise d’enfant sans vérif d’identité ni doublon',3,'P1','OUVERT'),
 ('clock.js','Horloge, jour d’école, rentrée, mode démo','3 tests','now() renvoie une Date (racine du bug NaN post-rechargement, 8+ modules) ; week-end dim/sam + semaine lun–ven EN DUR (BH = dim–jeu !) ; 08:00–15:00 fixe ; rentrée 15 sept fixe ; aucun fuseau',2,'P0','OUVERT — LE chantier calendrier'),
 ('currency.js','Devise active + money()','AUCUN direct','fr-FR en dur ; BHD sans millièmes (12,5 lu douze mille cinq cents)',2,'P0',FIX_REF + ' (décimales) ; locale à faire'),
 ('data.js','Emploi du temps enseignant + éval (questions/paniers)','AUCUN','SCHEDULE DE DÉMO EN DUR servi à TOUT enseignant ; teacherTimetable généré par hachage ≠ grille Direction ; jours/heures fixes',1,'P0','OUVERT'),
 ('db.js','Persistance : graine, migration v24, refs, liens parent↔enfant','8 tests','db() = parse COMPLET à chaque appel (chemins chauds par ligne !) ; install neuve = AUTO-GRAINE école démo + mots de passe publiés (Setup ne purge pas) ; quota 5 Mo ; uid Math.random',2,'P0','OUVERT — bloqueur d’expédition'),
 ('documents.js','Certificats numérotés, registre append-only','1 test','Numérotation length+1 (doublons après restauration) ; AUCUN gabarit rendu (le document physique n’existe pas) ; types tunisiens',3,'P1','OUVERT'),
 ('enrolment.js','Schéma multi-étapes pré-inscription','8 tests (le mieux couvert)','Placeholders tunisiens ; consentement générique au lieu du texte légal du pack ; pas de contrôle âge/niveau ; pas de détection de doublon',4,'P2','OUVERT'),
 ('facilities.js','Location d’installations','règle n°8','priceFor renvoie 0|objet → réservation à prix undefined → revenus NaN ; bookRecurring bug UTC ; encaissement HORS comptabilité',2,'P1','OUVERT'),
 ('features.js','Modules on/off + surcharges école','1 test','Surcharges lues à l’import (RN cassé) ; setModuleOverrides ne persiste pas ; pas de graphe de dépendances',3,'P2','OUVERT'),
 ('fetes.js','Fériés par pays + hégirien','4 tests','BH Fitr 2 j au lieu de 3 (corrigé) ; tables s’arrêtent à 2027 ; férié non relié présence/paie',3,'P1',FIX_REF + ' (Fitr 3 j)'),
 ('gallery.js','Moments photo, règle de vie privée dans le cœur','2 tests','consentOnly ignoré (corrigé) ; removeMoment suppression dure sans contrôle ; bug Date tri ; base64 dans le quota',2,'P0',FIX_REF + ' (consentOnly)'),
 ('hr.js','Contrats multi-composants, congés, paie maker-checker','4 tests','setBonus total périmé (corrigé) ; congés décomptent week-ends/fériés ; PAS de GOSI/EOSB (obligatoire BH) ; 2 modèles de congés parallèles',2,'P0',FIX_REF + ' (setBonus) ; GOSI/EOSB = CR-028'),
 ('i18n.js','FR→AR, direction, dates','3 tests','PAS d’anglais (BH = ar + en !) ; ~38 pages à 0 % ; t() sans interpolation/pluriels ; ar-TN pour BH',2,'P0','OUVERT — chantier majeur'),
 ('insights.js','Intelligence : climat, présence, recouvrement','4 tests','Climat vide sur données réelles (bug Date) ; recouvrement lit l’ANCIEN modèle payments (démo pour toujours)',3,'P1','OUVERT'),
 ('inventory.js','Stock léger avec mouvements','1 test','Journal tronqué à 30 mouvements (promesse de traçabilité rompue) ; pas d’unité/coût/péremption',3,'P2','OUVERT'),
 ('journal.js','Journal petite enfance (LE différenciateur)','AUCUN TEST','napMinutes NaN (bug Date) ; 2 parses par tape ; 2 tablettes s’écrasent ; pas de biberon/médicament/température ; pas de dé-envoi',2,'P0','OUVERT'),
 ('levels.js','Colonne vertébrale niveaux/cycles/modules','indirect i18n','École non configurée = PRIMAIRE par défaut → PERD la petite enfance (le différenciateur) ; libellés FR sans pack (KG1/Grade 1)',4,'P2','OUVERT'),
 ('livestatus.js','« Où est mon enfant » simulé','AUCUN','Été = juil–15 sept tunisien ; vendredi BH = affiche LUNDI « en classe » ; récrés en dur ; throw sur emploi du temps vide',2,'P0','OUVERT (chantier calendrier)'),
 ('locales.js','Packs pays BH/QA/TN/LY','12 tests (le pilier)','validId décoratif (accepte « abcde ») ; MANQUENT : week-end, fuseau, horaires, année scolaire, semestres, n° urgence — la racine de tous les bugs calendrier',4,'P1','OUVERT — étendre le pack'),
 ('mailer.js','Canal email unique','1 test','Pas de validation d’adresse ; échecs non persistés ; N envois non attendus par classe ; /api/mail INEXISTANT côté serveur',3,'P1','OUVERT'),
 ('mark.js / tokens.js / theme.js','Marque, jetons, accents par rôle','AUCUN','2 sources du logo ; FONT display Sora ≠ police chargée ; 2 hex inventés (hr/accountant) ; mix() sRGB naïf',4,'P3','OUVERT'),
 ('meteo.js','Météo Open-Meteo','AUCUN','Villes TUNISIENNES seulement — Manama voit la météo de TUNIS ; pas de timeout',2,'P1','OUVERT'),
 ('nav.js','Navigation étagée par rôle','2 tests i18n','NAV figée au chargement du module ; hr/accountant sans Tableau de bord/Messages/Annonces (corrigé)',3,'P2',FIX_REF),
 ('notify.js','Notifications in-app + email','2 tests','« École Al-Nour » signait chaque email (corrigé) ; markRead throw sur base neuve ; parents privés des diffusions par rôle en mode serveur ; croissance non bornée',2,'P1',FIX_REF + ' (nom école)'),
 ('oneroster.js','Export OneRoster v1.2 (argument de vente)','AUCUN TEST','Année 2025-2026 EN DUR ; classes à trait d’union perdent leurs inscriptions ; 200 classes synthétiques ; 1 seul enseignant/classe',2,'P2','OUVERT'),
 ('opsprofile.js','Fiche technique école','2 tests','Étiquettes live/démo INVERSÉES ; « Joignable : Oui » sans sonde ; 3 hôtes canoniques différents',3,'P2','OUVERT'),
 ('recruit.js','Pipeline recrutement','1 test','TypeError sur étape inconnue ; ids hors uid()/refs ; pas d’ACL interne ; PII candidats sans rétention',3,'P2','OUVERT'),
 ('refs.js','Référence ERP + uuid + QR','9 tests (excellent)','nextRef sous blob ACL réduit RÉÉMET la séquence 1 (doublons) ; route /verifier inexistante ; Luhn lettres faible',4,'P1','OUVERT'),
 ('requests.js','Circuit demandes : assigner→échéance→clore','2 tests','AUCUNE autorisation (chacun peut clore la demande d’autrui) ; monthReport mélange mois et aujourd’hui',3,'P1','OUVERT'),
 ('results.js','Scores, mentions, bulletins','1 test','mentionFor /100 EN DUR vs barème pays (/20 TN) → mentions FAUSSES ; bulletinFor O(jours×classes) et compte d’autres classes',2,'P0','OUVERT'),
 ('security.js','Poste de sécurité : urgences, consignes, rondes','AUCUN TEST','N° D’URGENCE TUNISIENS EN DUR (BH = 999) aussi dans le texte des consignes ; badge collision >999 ; FR seulement',1,'P0','OUVERT — sécurité des personnes'),
 ('social.js','Espaces & activités, quorum, double approbation','2 tests','sweep() mute pendant le rendu + parent ne persiste pas (états divergents) ; liste d’attente bloquée par une grande famille ; prix catalogues en DT implicite ; fuseaux',2,'P1','OUVERT'),
 ('storage.js','LA couture plateforme (web/RN/serveur)','1 test ciblé (excellent)','Plafond 5 Mo = LA limite produit (500 élèves ≈ 4,5 Mo) ; impl capturée au chargement',3,'P1','OUVERT (= hébergement backend)'),
 ('subjects.js','Matière → icône/teinte','AUCUN','Regex FR/latin : « English »/arabe → gris générique pour la moitié du programme du Golfe',3,'P2','OUVERT'),
 ('tunisia.js','Référentiel école : postes, docs, demandes','1 test délégation','« (DT) » en dur dans 3 libellés ; articles du code du travail TUNISIEN cités au BH ; hr/accountant ne peuvent déposer aucune demande',2,'P0','OUVERT'),
 ('workbench.js','« Ce qui attend MA décision »','4 tests (bien couvert)','hr/accountant/enseignant… SANS file (la RH ne voit pas la paie en attente !) ; deref chain sans garde peut blanchir le tableau de bord',4,'P1','OUVERT'),
 ('remote.js (app)','La couture client↔serveur','AUCUN TEST','pushing bloqué à jamais à la 1re coupure (corrigé) ; 409 = écrasement local (copie de sauvegarde ajoutée) ; blob entier à chaque push ; jeton en localStorage',2,'P0',FIX_REF + ' (try/finally + copie)'),
 ('RemoteGate.jsx (app)','Porte du mode serveur','AUCUN','2e écran de connexion non brandé FR-seulement ; PAS de « mot de passe oublié » (le vrai mode client !) ; jeton expiré = boucle de rechargement',2,'P0','OUVERT'),
]
ws = sheet(wb, 'Modules (cœur)', 'MODULES DU CŒUR (core/src)', f'51 modules · audit {AUDIT_DATE} · trié par nom', ['Module','Rôle','Tests','Manques / bugs clés','Prépa 1-5','Priorité','Statut'], [22,34,26,66,9,9,26])
fill(ws, M, sev_col=6)

# ═══ 3. PAGES WEB ═════════════════════════════════════════════════════════════
P = [  # (page, route, rôles, constat, manque école réelle, prep, prio, statut)
 ('Login','/login','public','Identifiants démo EN CLAIR + connexion 1-clic tous rôles (owner incl.) expédiés en production','Logo école, SSO, 1re connexion → changement de mot de passe',2,'P0','ATTÉNUÉ (redirigé en mode serveur) ; à gater par build'),
 ('Dashboard','/app','9 rôles','7 tableaux en 1 fichier (579 l.) ; crash si SCHEDULE vide ; hex hors jetons','Sélecteur de période, export KPI',3,'P1','OUVERT'),
 ('Students','/app/students','direction/admin/surv./ens.','Défauts Tunisie (nationalité, gouvernorat, 5ème année, loi 2004-63) ; AUCUNE édition/archivage ; notification groupée sans confirmation','Édition, transfert de classe, IMPORT EN MASSE, photo, fratries, badge',3,'P0','Import → CR-034 (en cours)'),
 ('StudentProfile','/app/eleve/:id','direction/admin/surv./ens.','100 % FR ; lecture seule totale ; « Prévenir le parent » sans confirmation','Édition, impression dossier, journal des consultations',3,'P1','OUVERT'),
 ('Teachers','/app/teachers','direction/admin','DT en dur, CIN/Gouvernorat en dur, pas d’édition, mur de cartes sans pagination, salaire visible par admin','Édition, affectation classes/matières, lien contrat RH',2,'P0','OUVERT'),
 ('Accounts','/app/accounts','direction','Nouveau mot de passe AFFICHÉ EN TOAST ; désactivation sans confirmation ; « Identité (Tunisie) » ; ChildPicker inutilisable >80 élèves','Création en masse, invitation par email, politique de mot de passe, dernière connexion',3,'P0','OUVERT'),
 ('Attendance','/app/attendance','direction/ens./surv./admin','Aujourd’hui SEULEMENT (pas de rattrapage) ; défaut = tous présents sauvés en silence ; notifications sans confirmation ni dédoublonnage','Navigation par date, codes d’absence justifiée, registre mensuel exportable (exigence GCC)',3,'P0','OUVERT'),
 ('Evaluate','/app/evaluate','enseignant','Questions figées dans le cœur ; brouillon partagé entre onglets ; pas d’édition après envoi','Rubrique configurable, historique enseignant',3,'P1','OUVERT'),
 ('Results','/app/results','direction/admin','« Top élèves » CLASSE les enfants (contredit la règle n°9 du produit) ; seuils en dur ; pas d’export','Comparaison de périodes, rapport imprimable',3,'P2','OUVERT — décision produit à prendre'),
 ('Academic','/app/academic','direction/admin/ens.','Pas de garde « modifications non enregistrées » ; « Exécuter quand même » aussi visible que Confirmer (irréversible + argent) ; /20 français','PDF bulletin ici, pondération des périodes, signature parent',3,'P1','OUVERT'),
 ('Accounting','/app/accounting','direction/comptable','« Facturer tous les élèves » SANS confirmation ni aperçu (factures immuables !) ; liste sans recherche/filtre/pagination ; remise révoquée d’un clic','Échéanciers, impression de reçu, relevé impayés, rapprochement, TVA',3,'P0','OUVERT'),
 ('Finance','/app/finance','direction/comptable','Grille couleur-seulement (daltonisme/tactile) ; clic = bascule immédiate (un raté « décaisse » et notifie le parent) ; crash si élève sans échéancier ; DOUBLE VÉRITÉ avec Accounting','Montants affichés, vue par famille, paiements partiels',2,'P0','OUVERT'),
 ('Payments (parent)','/app/payments','parent','AUCUN MONTANT nulle part ; pas de reçu ; déclaration groupée sans confirmation','Paiement en ligne (Benefit BH !), facture PDF, échéances',2,'P0','OUVERT — Tier-1 Golfe'),
 ('Admissions','/app/admissions','direction/admin','Tuiles inertes (rompt « chaque chiffre s’ouvre ») ; refus instantané sans confirmation (email !) ; pas de recherche','Entretiens, position liste d’attente, frais de dossier, doublons',3,'P1','OUVERT'),
 ('Inscription (public)','/inscription','PUBLIC','Sans CAPTCHA ni limite (démo) ; placeholders tunisiens ; pas de brouillon ; pas de relecture avant envoi','Anti-spam, vérification email/tél., suivi par référence, multi-enfants',3,'P0','OUVERT (limite serveur 10/h existe)'),
 ('Setup','/setup','était PUBLIC','N’importe qui réécrivait nom + niveaux de l’école','Pays/devise/langue/calendrier au premier lancement, purge de la démo',2,'P0',FIX_REF + ' (protégé) ; assistant à enrichir'),
 ('Settings','/app/settings','direction','Pas de garde sortie ; reload à 750 ms ; pas de week-end/fuseau/horaires/calendrier','Semaine dim–jeu (BH !), année scolaire, logo, sauvegarde/export complet',3,'P0','OUVERT'),
 ('HR','/app/hr','direction/RH','Maker-checker solide ; MAIS prime = input silencieux onBlur ; « Marquer payée » irréversible sans confirmation ; fr-FR','GOSI/impôts, EOSB, export WPS/banque, historique',3,'P1','OUVERT (CR-028 suite)'),
 ('Staff','/app/staff','direction/RH','isWeekend() sam/dim → décomptes FAUX au BH ; QUOTA 30 en dur ; DOUBLE système de congés vs HR ; listes de personnel divergentes','Shifts, heures sup., import biométrique',2,'P0','OUVERT (chantier calendrier)'),
 ('Pointage','/app/pointage','ens./surv./sécu/admin','Pointage depuis n’importe où ; re-pointage EFFACE la journée ; seuils en dur','Horaires configurables, correction demandée, approbation',3,'P1','OUVERT'),
 ('Recruit','/app/recruit','direction/RH','Fermer/rouvrir sans confirmation ; pas de CV joint (Attach existe !)','Entretiens, lettres, rétention RGPD candidats',3,'P2','OUVERT'),
 ('Journal','/app/journal','ens./admin/direction/parent','Aujourd’hui seulement ; tape = écrit sans annulation ; note perdue si onglet fermé','Historique par enfant, photos, médicaments, PDF semaine',3,'P1','OUVERT'),
 ('ChildFile','/app/child','ens./admin/direction','Révocation de personne autorisée SANS confirmation (la liste la plus critique du produit) ; remise d’enfant 1-clic ; CIN en dur','Photo de la personne autorisée, autorisation temporaire, garde parentale',3,'P0','OUVERT'),
 ('Behavior','/app/behavior','ens./admin/direction/parent','Score numérique par enfant visible (signal de classement) ; suppression sans confirmation','Intervention/escalade, accusé parent',3,'P2','OUVERT'),
 ('Gallery','/app/gallery','ens./admin/direction/parent','confirm() natif ; base64 → quota ; PAS de modèle de consentement photo à la prise','Consentement par enfant appliqué au partage, vidéo, albums, modération',2,'P0','consentOnly cœur corrigé ; UI consentement OUVERTE'),
 ('Canteen','/app/canteen','ens./admin/direction/parent','Alerte allergie = texte libre FR (voir canteen.js) ; semaine figée sans navigation','Champ allergène structuré, effectifs cuisine, halal/régimes',3,'P1','OUVERT'),
 ('Accidents','/app/accidents','ens./surv./admin/direction/parent','Meilleur flux du produit ; envoi au parent sans confirmation ; tuiles inertes ; parent arabophone reçoit un document FR','PDF assureur, transfert hôpital, bouton urgence',4,'P1','OUVERT'),
 ('Documents','/app/documents','admin/direction','« , Tunisie » DANS le PDF ; helvetica → l’ARABE NE S’IMPRIME PAS ; Imprimer imprime toute l’app','Logo/en-tête, gabarit arabe/bilingue, cachet',2,'P0','OUVERT'),
 ('Budget','/app/budget','direction/comptable','prompt() natif pour le motif d’annulation (piste d’audit !) ; pas de prévisionnel','Lignes budgétaires, centres de coût, seuils d’approbation',3,'P1','OUVERT'),
 ('Inventory','/app/inventory','direction/admin','±1 par clic (500 feuilles = 500 clics) ; pas d’édition/suppression ; note de mouvement jamais saisie','Quantité saisissable, coût, code-barres, inventaire tournant',3,'P2','OUVERT'),
 ('Facilities','/app/facilities','direction/admin','Encaissement forcé « espèces » sans confirmation ni reçu ; hours casse sur demi-heures ; placeholder tunisien','Contrat PDF, caution, calendrier, clients externes',3,'P2','OUVERT'),
 ('Security','/app/security','sécu/direction/admin','N° d’urgence TUNISIENS (vie des personnes !) ; ronde en cours perdue au refresh ; pas de recherche registre','N° par pays, badge visiteur imprimé, appel de rassemblement',3,'P0','OUVERT'),
 ('Social','/app/social','6 rôles','money = n+" DT" EN DUR ; sweep() mute au chargement ; 609 l. FR ; pas de lien vers la compta pour l’argent collecté','Encaissement réel, présence le jour J',3,'P1','OUVERT'),
 ('Events','/app/events','6 rôles','Pas d’édition (supprimer+recréer) ; corbeille invisible au tactile ; hr/accountant/sécu non ciblables ; semaine lundi','Récurrence, import calendrier officiel, RSVP',3,'P1','OUVERT'),
 ('Incidents','/app/incidents','surv./sécu/admin/direction','55 lignes ; échec silencieux si titre vide ; résolution sans note ; email parent automatique sans aperçu ; 3 systèmes d’incidents se chevauchent','Notes de résolution, pièces jointes, escalade',2,'P0','OUVERT'),
 ('Requests','/app/requests','ens./admin/direction/RH','PDF « Tunis, le… , Tunisie, DT » + LOGO ÉTRANGER en SVG inline ; helvetica/arabe ; état comment réutilisé 3 fois','Types configurables par école, pièces jointes, SLA',2,'P0','OUVERT'),
 ('Messages','/app/messages','9 rôles','JAMAIS marqué lu ; pas de diffusion de groupe (25 parents = 25 messages) ; pas de pièces jointes','Diffusion classe, accusés, rétention légale',2,'P0','OUVERT'),
 ('Notices','/app/notices','9 rôles','Stockées comme notifications (pas d’entité) : ni édition ni expiration ni pièce jointe ; admin voit TOUT y compris ciblé','Riche texte, circulaires PDF, publication planifiée, accusés',2,'P0','OUVERT'),
 ('Notifications','/app/notifications','9 rôles','inboxFor ×5 par rendu ; boîte infinie sans purge','Préférences par canal/type, résumé email',3,'P2','OUVERT'),
 ('Live','/app/live','parent','SIMULÉ depuis l’emploi du temps mais badge « EN DIRECT » pendant les heures de classe (confiance !) ; 08:00–15:00 en dur','Pointage portail réel (RFID/QR) pour rendre le direct vrai',3,'P2','OUVERT — honnêteté à corriger'),
 ('Timetable','/app/timetable','5 rôles','AUCUNE détection de conflit (enseignant/salle en double) ; salles = 8 chaînes en dur ; lun–ven','Conflits, remplacements, impression, périodes par pays',2,'P0','OUVERT'),
 ('Homework/Exams/Library/Transport','4 routes','modules ÉTEINTS','Squelettes ; Transport a un faux « EN DIRECT » calculé de l’horloge (danger de confiance si activé)','Ne PAS activer en l’état',1,'P3','OUVERT — documentés éteints'),
 ('Schools (owner)','/app/schools','owner','MOT DE PASSE ADMIN EN CLAIR à l’écran ; prix DT en dur ; « identifiants envoyés » = mensonge (rien n’est envoyé)','Provisionnement réel, facturation, métriques',2,'P0','OUVERT'),
 ('Interop','/app/interop','direction/admin','8 téléchargements en rafale (bloqués par navigateur) ; busy jamais rendu ; export seul','ZIP, IMPORT (→ CR-034), synchronisation planifiée',4,'P2','Import → CR-034'),
 ('Site public + Pricing','/, /modules, /tarifs…','public','€ sur /tarifs vs DT dans la console owner ; « €79 € » double symbole ; pas d’auto-inscription','Tarifs BHD/région, parcours contact-vente',3,'P2','OUVERT'),
 ('Landing.jsx','AUCUNE ROUTE','—','443 lignes ORPHELINES (remplacée par site/*) — à supprimer','—',0,'P2','OUVERT'),
 ('Composants partagés','ui/DataTable/Bulletin/Payslip/Attach…','—','DataTable (excellent) utilisé par UNE page ; ErrorState/Skeletons jamais utilisés (aucun état chargement/erreur) ; Bulletin « République Tunisienne » ; Payslip sans cotisations ; AUCUN error boundary → écran blanc','Adopter DataTable partout ; error boundary global ; gabarits par pays',3,'P0','OUVERT'),
]
ws = sheet(wb, 'Pages Web', 'PAGES WEB (app/src/pages)', f'~50 pages · audit {AUDIT_DATE}', ['Page','Route','Rôles','Constats (UX/fonctionnel)','Ce qu’une vraie école attend','Prépa','Priorité','Statut'], [16,14,16,52,38,7,8,22])
fill(ws, P, sev_col=7)

# ═══ 4. ÉCRANS MOBILE ═════════════════════════════════════════════════════════
MOB = [
 ('Welcome / Login','pré-auth','Rituel de marque + connexion (pastilles démo sans hr/comptable/owner)',4,'OUVERT'),
 ('Dashboard','tous','130 l. vs 579 web : sans fêtes/météo/intelligence/graphiques ; tuile Annonces lisait une collection inexistante',3,FIX_REF + ' (tuile)'),
 ('Live / Bulletin / Social / Security / Requests / Evaluate / Attendance / Events / Incidents / Payments','divers','Les 10 écrans « pleins » — au niveau du web',5,'OUVERT (i18n)'),
 ('Finance / Pointage / Results / Students / Teachers / Timetable / Messages / Notices / Accidents / Canteen','divers','Partiels : lecture seule ou sans édition/pièces jointes',4,'OUVERT'),
 ('Notifications','tous','ORPHELIN : dans le routeur mais pas dans NAV — inatteignable pour le personnel ; aucune cloche/badge dans le Shell',3,'OUVERT'),
 ('Staff / Behavior / Journal / Gallery','divers','Lecture seule (bannière « saisie sur le web ») — les 4 différenciateurs petite enfance NON saisissables au téléphone',3,'OUVERT'),
 ('Shell.js','—','switchTab sans canAccess (contournement de rôle)',4,FIX_REF),
 ('App.js (boot)','—','Pack pays/devise/programme jamais initialisés → Bahreïn voyait dinars et matières tunisiens',3,FIX_REF),
]
ws = sheet(wb, 'Écrans Mobile', 'ÉCRANS MOBILE (mobile/src/screens)', '28 écrans · Expo SDK 54 / RN 0.81', ['Écran(s)','Rôles','Constat','Prépa','Statut'], [34,14,72,7,24])
fill(ws, MOB)

# ═══ 5. PARITÉ WEB↔MOBILE ═════════════════════════════════════════════════════
PAR = [
 ('Arabe / RTL / i18n','complet (shell+parent ~25-30 %)','MANQUANT TOTAL : zéro t(), zéro I18nManager, police latine seule','P0 — inutilisable à Manama'),
 ('Mode serveur (école réelle)','existe (remote.js, web-only)','MANQUANT : le téléphone ne voit QUE la démo locale','P0 — les parents verraient des enfants fictifs'),
 ('Notifications push','n/a (in-app)','MANQUANT : expo-notifications absent — la raison n°1 d’installer l’app','P0'),
 ('Pack pays (devise, pièces, programme)','complet','Initialisé au boot désormais ; écrans Social/Teachers gardent « DT » en dur','P1 — ' + FIX_REF + ' (boot)'),
 ('Saisie petite enfance (journal, comportement, photos, accidents)','complet','LECTURE SEULE mobile — les différenciateurs ne se saisissent pas au téléphone','P0 pour les enseignants'),
 ('Récupération de mot de passe','complet (web)','MANQUANT mobile','P1'),
 ('11 modules direction (Admissions, Compta, Académique, RH+bulletin de paie, Budget, Documents, Installations, Inventaire, Recrutement, Interop, Paramètres/Comptes)','complets','MANQUANTS → cartes « Bientôt »','P1-P2 (assumé, à documenter au client)'),
 ('Fêtes/hégirien, météo, palette Ctrl+K','complets','MANQUANTS mobile','P2-P3'),
 ('Robustesse mobile','—','Aucune FlatList (120 élèves en ScrollView), aucun error boundary par écran, aucune accessibilité, writes AsyncStorage silencieux, eas.json squelette, dist/ commité','P1'),
 ('Test mobile','18 parcours web','1 script (3 rôles, react-native-web, assertions faibles)','P1'),
]
ws = sheet(wb, 'Parité Web-Mobile', 'PARITÉ WEB ↔ MOBILE', 'Exigence produit : les deux plateformes synchrones', ['Capacité','Web','Mobile','Verdict'], [40,26,52,26])
fill(ws, PAR)

# ═══ 6. API SERVEUR ═══════════════════════════════════════════════════════════
API = [
 ('GET/ANY','/api/health','aucune','—','aucune','Fuit la révision (oracle d’activité) publiquement — à restreindre'),
 ('POST','/api/login','aucune','—','20/15 min/IP','Limite AJOUTÉE (' + FIX_REF + ') ; scrypt N=16384 < plancher OWASP ; bloque l’event loop'),
 ('POST','/api/logout','jeton','—','aucune','OK'),
 ('POST','/api/forgot','aucune','—','5/h/IP','Réponse neutre constante — bon design'),
 ('POST','/api/reset','jeton reset','—','10/h/IP','Limite AJOUTÉE (' + FIX_REF + ') ; jeton 192 bits usage unique 60 min'),
 ('POST','/api/apply','aucune (public)','—','10/h/IP','…mais ...rest non validé côté serveur : peut écraser id/stage — À CORRIGER ; 30 Mo de corps accepté'),
 ('GET','/api/db','session','parent→blob reconstruit ; staff→strip par rôle','aucune','READ_STRIP.admin=[] → admin lit paie/argent — À CORRIGER'),
 ('POST','/api/db','session','parent 403 ; mergeWrite par rôle','aucune','Boucle mots de passe GATÉE par ACL (' + FIX_REF + ') ; verrou de révision 409 prouvé sous 20 écritures concurrentes ; « * » remplace en bloc (blob tronqué = collections effacées) — À CORRIGER'),
 ('POST','/api/op','session parent','acknowledge / toggleLike gardés par childIds','aucune','OK'),
 ('—','/api/mail','—','—','—','INEXISTANT alors que remote.js l’appelle → tout email en mode serveur échoue — À CRÉER'),
 ('GET','/* (statique)','aucune','—','aucune','Traversée de chemin CORRIGÉE (' + FIX_REF + ') + nosniff ; CSP/HSTS restent à poser au reverse-proxy'),
 ('(transverse)','sauvegardes','—','—','—','auth.json désormais SAUVEGARDÉ + 0600 (' + FIX_REF + ') ; copie hors machine + restauration RÉPÉTÉE = jour J hébergement'),
 ('(transverse)','journal d’audit','—','—','—','ABSENT — exigence PDPL/INPDP (« qui a lu le dossier médical ») ; n°1 du gap-analysis §4'),
]
ws = sheet(wb, 'API Serveur', 'API DU SERVEUR (server/server.mjs)', 'Un processus, une école (pilote v1) · zéro dépendance', ['Méthode','Chemin','Auth','Contrôle de rôle','Limite','Notes'], [10,16,14,26,12,60])
fill(ws, API)

# ═══ 7. BUGS & RISQUES ════════════════════════════════════════════════════════
B = [
 ('S-1','CRITIQUE','server/server.mjs:260','Boucle mots-de-passe AVANT l’ACL : tout rôle authentifié remplaçait l’identifiant de la Direction (prise de contrôle totale)',FIX_REF),
 ('S-2','CRITIQUE','server/server.mjs:295','Traversée de chemin du statique : /../data/auth.json servi (hachages + sessions)',FIX_REF),
 ('S-3','CRITIQUE','app/src/components/Attach.jsx:117','XSS stocké via le NOM de fichier du formulaire public, exécuté dans la session admin (vol de jeton)',FIX_REF),
 ('S-4','HAUTE','app/src/mail.js:15','Jeton du worker mail commité dans un dépôt PUBLIC (spam via contact@)','RÉVOQUÉ + worker durci (Origin exigé, débit limité) + secret CI — ' + FIX_REF),
 ('CC-1','CRITIQUE','core/src/clock.js (now())','now() renvoie une Date → chaîne après rechargement → NaN dans accidents/comportement/journal/gallery/insights/HR sur DONNÉES RÉELLES (invisible en CI)','OUVERT — chantier n°1 du cœur'),
 ('CC-2','CRITIQUE','core/src/db.js:726','Install neuve = AUTO-GRAINE de l’école démo tunisienne (120 faux élèves, mots de passe publiés) ; Setup ne purge jamais','OUVERT — bloqueur d’expédition'),
 ('B-2','CRITIQUE','clock/data/db/livestatus/Staff…','Semaine lun–ven et week-end sam-dim EN DUR : au BH (dim–jeu) présence, emploi du temps, direct, cantine et PAIE tombent les mauvais jours','OUVERT — étendre locales.js (weekend/weekStart) puis router tout getDay()'),
 ('B-3','CRITIQUE','core/src/accounting.js','AUCUNE TVA (BH 10 % + e-invoicing NBR) : pas de facture fiscale conforme possible','OUVERT — CR à ouvrir'),
 ('S-6','HAUTE','core/src/admissions.js:95','/api/apply : ...rest non validé peut écraser id/stage ; fichiers sans contrôle serveur','OUVERT'),
 ('S-7/8','HAUTE','core/auth.js·accounts.js','Mots de passe en clair (démo), défaut 1234, reset Math.random 6 car.','OUVERT (mode démo) — politique à poser au passage serveur'),
 ('ACL-1','HAUTE','core/src/acl.js:62','READ_STRIP.admin=[] : l’Administration LIT paie/factures/salaires','OUVERT'),
 ('ACL-2','HAUTE','core/src/acl.js:156','mergeWrite « * » : un blob direction tronqué EFFACE des collections serveur ; users sans pw écrasent le registre','OUVERT'),
 ('HR-1','CRITIQUE','core/src/hr.js:204','setBonus : total de paie périmé validé/payé/comptabilisé',FIX_REF),
 ('GAL-1','HAUTE','core/src/gallery.js:79','consentOnly stocké mais jamais appliqué (PDPL)',FIX_REF),
 ('CAN-1','CRITIQUE','core/src/canteen.js:33','Alerte allergie : correspondance FR uniquement — dossier en anglais/arabe = AUCUNE alerte (sécurité enfant)','OUVERT'),
 ('SEC-1','CRITIQUE','core/src/security.js:21','Numéros d’urgence TUNISIENS en dur (BH = 999) affichés au gardien pendant un incendie','OUVERT — champ urgences dans le pack pays'),
 ('RES-1','CRITIQUE','core/src/results.js:34','mentionFor /100 en dur vs barème pays → mentions fausses sur bulletins','OUVERT'),
 ('DATA-1','CRITIQUE','core/src/data.js:29','Emploi du temps DE DÉMO servi à tout enseignant ; grille enseignant générée par hachage ≠ grille Direction','OUVERT'),
 ('REM-1','CRITIQUE','app/src/remote.js:106','pushing bloqué à jamais à la 1re coupure réseau (perte silencieuse d’une journée)',FIX_REF),
 ('REM-2','HAUTE','app/src/remote.js:112','409 = écrasement local sans merge','Copie locale coreon_db_conflict ajoutée (' + FIX_REF + ') ; merge réel À FAIRE'),
 ('SET-1','HAUTE','app/src/App.jsx:111','/setup public réécrivait l’école',FIX_REF),
 ('S-9','HAUTE','app/src/pages/Login.jsx','loginAs 1-clic joignable en mode serveur','Redirigé (' + FIX_REF + ') ; à retirer du build client'),
 ('ACC-1','HAUTE','core/src/accounts.js:20','hr/accountant non créables par le client',FIX_REF),
 ('NAV-1','MOYENNE','core/src/nav.js','hr/accountant sans Tableau de bord/Messages/Annonces',FIX_REF),
 ('MOB-1','HAUTE','mobile/src/Shell.js:70','switchTab sans canAccess (contournement de rôle)',FIX_REF),
 ('MOB-2','MOYENNE','mobile/screens/Dashboard.js:100','Tuile Annonces lisait d.notices (inexistant) → 0 pour toujours',FIX_REF),
 ('MOB-3','CRITIQUE','mobile/App.js','Pack pays jamais initialisé (BH voyait DT/matières tunisiennes)',FIX_REF),
 ('NOT-1','HAUTE','core/src/notify.js:30','« École Al-Nour » signait chaque email de chaque école',FIX_REF),
 ('CUR-1','CRITIQUE','core/src/currency.js:8','BHD sans millièmes : 12,5 lu « douze mille cinq cents » sur factures/paie',FIX_REF),
 ('FET-1','MOYENNE','core/src/fetes.js','Aïd el-Fitr BH : 2 j au lieu de 3',FIX_REF),
 ('REF-1','HAUTE','core/src/refs.js:137','nextRef sous blob ACL réduit réémet la séquence 1 (références en double)','OUVERT'),
 ('NUM-1','HAUTE','accounting/documents nextNumber','Séries = length+1 : réémission de numéros après restauration/fusion (faute d’audit)','OUVERT'),
 ('MAIL-1','HAUTE','server + remote.js','/api/mail appelé par le client, jamais implémenté → zéro email en mode serveur','OUVERT'),
 ('CI-1','MOYENNE','.github/workflows/deploy.yml','npm audit ne bloquait jamais ; lint annoncé jamais exécuté','Audit CRITIQUE bloque + oxlint ajouté — ' + FIX_REF),
 ('CI-2','MOYENNE','e2e/','~340 attentes fixes, ports en collision, 14/15 diags sans code de sortie, smoke.production.mjs (le plus fort) HORS CI','OUVERT'),
 ('TST-1','MOYENNE','core/test/core.test.mjs','1 fichier, base singleton, 13/100 resets → ordre-dépendant ; dates épinglées 2026-2027','OUVERT'),
 ('STO-1','HAUTE','core/src/storage.js','Plafond localStorage ~5 Mo : 500 élèves ≈ 4,5 Mo SANS photos — la limite produit','OUVERT = hébergement backend (T-EDU-BACKEND)'),
 ('AUD-1','CRITIQUE','transverse','AUCUN journal d’audit (qui a lu/modifié quoi) — exigence PDPL/INPDP, n°1 entreprise','OUVERT — après CR-029'),
]
ws = sheet(wb, 'Bugs & Risques', 'REGISTRE DES BUGS & RISQUES', f'Consolidé des 6 audits du {AUDIT_DATE} · filtrer par Sévérité/Statut', ['ID','Sévérité','Où','Description','Statut'], [8,11,26,66,30])
fill(ws, B, sev_col=2)

# ═══ 8. BAHREÏN GO-LIVE ═══════════════════════════════════════════════════════
BH = [
 (1,'Semaine dim–jeu / week-end ven-sam','Présence, emploi du temps, direct, cantine, décomptes de PAIE tombent les mauvais jours','Étendre locales.js (weekend, weekStart, schoolDays) + router clock/data/livestatus/Staff/Events par le pack','OUVERT — le plus gros défaut fonctionnel'),
 (2,'Chantier now() → nombre','Climat, accidents, journal, moments cassés sur données réelles','now() epoch ms + normalisation en lecture + tests','OUVERT'),
 (3,'Purge de la démo à l’installation','Le client démarre avec 120 faux élèves tunisiens et des mots de passe publiés','Setup : pays/langue/devise/purge + création du 1er compte direction','OUVERT'),
 (4,'TVA 10 % + facture fiscale NBR','La finance ne signe pas sans facture conforme','taxRate/taxNumber au pack + lignes de taxe + gabarit facture','OUVERT'),
 (5,'Alerte allergie multilingue + champ structuré','Sécurité enfant : dossiers EN/AR = aucune alerte','Table d’allergènes multilingue + champ structuré au dossier','OUVERT'),
 (6,'Urgences par pays (999) + consignes','Vie des personnes ; affiché au gardien','Champ urgences dans le pack + Security.jsx/security.js le lisent','OUVERT'),
 (7,'Mentions par barème pays (results.js)','Bulletins faux en /100 vs /20','mentionFor → locales.gradeOf','OUVERT'),
 (8,'CPR partout où l’écran dit CIN','Accounts, ChildFile (portail !), Teachers, Security, Requests','Appeler idLabelFor()/idTypesFor() déjà existants','OUVERT'),
 (9,'Anglais (3e locale) + AR des 38 pages métier','BH = arabe + anglais ; l’app métier est FR-seulement à ~70 %','Locale en + dictionnaire ; prioriser les écrans du personnel','OUVERT — chantier majeur'),
 (10,'PDF/impression arabe (helvetica → carrés)','Bulletins, certificats, demandes illisibles en arabe','Police arabe embarquée jsPDF + gabarits bilingues + logo école','OUVERT'),
 (11,'Paiement en ligne Benefit','Déclencheur d’achat n°1 du Golfe (Tier-1)','Après hébergement backend ; passerelle par pays','PLANIFIÉ (gap-analysis §1)'),
 (12,'Hébergement backend + HTTPS + restauration testée','Sans serveur : localStorage, mots de passe en clair, 5 Mo','T-EDU-BACKEND (Workers Option 0 vs Hetzner)','OUVERT — LA priorité structurelle'),
 (13,'GOSI/EOSB dans la paie','Obligation légale BH ; bulletin sans cotisations non utilisable','CR-028 volets suivants','PLANIFIÉ'),
 (14,'Niveaux KG1/KG2/Primary + 2 semestres','Nommage MoE Bahreïn','levels + TERMS dans le pack pays','OUVERT'),
 (15,'Positifs déjà en place','Hégirien natif, fériés BH (3 j Fitr corrigé), consentement PDPL, programme /100 A-F, CPR défini au pack, devise à 3 décimales','—','ACQUIS'),
]
ws = sheet(wb, 'Bahreïn Go-Live', 'BAHREÏN — CE QUI BLOQUE LE 1ER CLIENT', 'Classé par ordre d’exécution recommandé', ['#','Chantier','Pourquoi','Comment','Statut'], [4,30,42,42,26])
fill(ws, BH)

# ═══ 9. TESTS & CI ════════════════════════════════════════════════════════════
T = [
 ('Cœur : 100 tests','1 fichier, singleton db, 13 resets → ordre-dépendant ; dates épinglées 2026-27 ; certains lisent l’horloge réelle','Reset par test ou fixtures ; geler l’horloge ; découper le fichier','P1'),
 ('Modules SANS test unitaire','charts, currency, data, journal (différenciateur !), livestatus, mark, meteo, oneroster (argument de vente !), security, subjects, theme, tokens','Écrire les tests journal + oneroster + security d’abord','P1'),
 ('Parcours e2e : 18 en CI','~340 waitForTimeout, ports en collision (6 paires), serveDist répond 200 à tout (404 invisibles), pas d’artefacts d’échec','waitForSelector, listen(0), captures/traces sur échec, timeout-minutes','P1'),
 ('Suites HORS CI','smoke.production.mjs (RBAC/négatif — le plus rigoureux du dépôt), 15 diag.* dont 14 sans code de sortie (pays/Bahreïn incl.)','Codes de sortie partout + intégrer au moins smoke.production + diags pays en CI','P0'),
 ('Parcours manquants (test-plan.md)','Chaîne accident UI bout-en-bout ; paie préparer→valider→verrouiller ; fuite notifications inter-familles (régression réelle 2026-07-10)','Écrire les 3 parcours','P1'),
 ('Serveur : 15 tests','Verrou de révision prouvé sous 20 écritures ; MAIS compteurs de débit partagés entre tests (ordre-dépendant)','Réinitialiser les limiteurs par test','P2'),
 ('Mobile : 1 script','3 rôles, react-native-web, assertions faibles (login + 4 écrans + zéro pageerror)','Maestro/Detox sur build réel ; étendre rôles','P1'),
 ('CI deploy.yml','Audit ne bloquait pas (corrigé : critique bloque) ; lint absent (corrigé) ; navigateurs non cachés ; concurrency PR annule un déploiement main ; aucune notification d’échec','Cache navigateurs ; groupes de concurrency séparés ; notification (email/Telegram)','P1'),
 ('Environnements','dev → dev.edu (rapide) · int → int.edu (chaîne complète) · main → prod. En place depuis le ' + AUDIT_DATE,'Discipline de flux dev→int→main + tags v1.x.y','FAIT (CR-031)'),
 ('Rollback','workflow_dispatch depuis SHA sain, ou revert ; smoke-prod détecte sans agir','Documenter + répéter un rollback réel une fois','P2'),
]
ws = sheet(wb, 'Tests & CI', 'TESTS & CHAÎNE CI/CD', 'État et trous de la preuve de qualité', ['Zone','Constat','Action','Priorité'], [26,58,40,10])
fill(ws, T, sev_col=4)

# ═══ 10. DETTES & DÉCISIONS ══════════════════════════════════════════════════
D = [
 ('Journal d’audit immuable','n°1 du gap-analysis §4 ; exigence PDPL/INPDP ; « qui a changé cette note ? »','Après CR-029 (ordre D-décisions)','P0 structurel'),
 ('Isolation multi-tenant','Serveur = une école par processus ; préalable à la 2e école','Avec l’hébergement','P1'),
 ('Hébergement backend (T-EDU-BACKEND)','Tout le reste en découle : mots de passe hachés, quota, multi-appareils, emails serveur','Workers (option 0, sans carte) vs Hetzner — évaluer S1','P0 structurel'),
 ('Double vérité Finance vs Accounting','2 modèles d’argent (payments[] vs invoices) sans lien','Unifier sur invoices/receipts ; migrer la grille','P1'),
 ('Double système de congés (Staff vs HR)','2 soldes divergents, 2 listes de personnel','Unifier sur hr.js','P1'),
 ('3 systèmes d’incidents (Incidents/Accidents/Behavior)','Chevauchement, confusion de saisie','Clarifier les périmètres, lier les dossiers','P2'),
 ('DataTable partout','Le seul composant liste digne (recherche/tri/pagination/CSV) est utilisé par 1 page sur ~20 listes','Adoption progressive, Accounts/Teachers/Accounting d’abord','P1'),
 ('Error boundary + états chargement','Aucun : tout throw = écran blanc ; ErrorState/Skeletons existent inutilisés','Boundary global + par page lazy','P0 rapide'),
 ('Confirmations des actions irréversibles','~15 actions d’argent/sécurité partent d’un clic sec (liste en feuille Pages)','Composant confirm() maison systématique','P0 rapide'),
 ('HANDOFF.md périmé','Recommande encore Supabase, ignore server/','Réécrire ou supprimer','P2'),
 ('Chiffres de tests incohérents dans les docs','README « 19 tests » vs 100 réels ; « 276/276 »','Générer les chiffres, ne plus les écrire à la main','P3'),
 ('Landing.jsx orpheline (443 l.)','Code mort qui divergera','Supprimer','P2'),
 ('« Top élèves » vs règle n°9','Le produit s’interdit de classer les enfants… et le fait dans Results','DÉCISION OTHMAN : retirer ou assumer ?','DÉCISION'),
 ('Badge « EN DIRECT » simulé (Live/Transport)','Confiance parent : le direct est un emploi du temps','DÉCISION : renommer « journée type » tant que pas de pointage réel ?','DÉCISION'),
 ('Identifiants démo sur /login public','Vitrine commerciale vs sérieux production','DÉCISION : build « client » sans démo vs garder la vitrine sur edu.*','DÉCISION'),
]
ws = sheet(wb, 'Dettes & Décisions', 'DETTES STRUCTURELLES & DÉCISIONS À PRENDRE', 'Ce qui ne se corrige pas en un commit', ['Sujet','Constat','Voie','Priorité'], [30,52,40,14])
fill(ws, D)

# ═══ 11. FAT EXPLORATOIRE 2026-07-26 (5 rapports QA) ═════════════════════════
FAT = [
 ('REJET','Vérité UI','« Suspendre l\'école » ne coupait AUCUN accès (badge décoratif)', QA_REF),
 ('REJET','Vérité UI','Le provisionnement affichait un mot de passe sans CRÉER le compte', QA_REF),
 ('REJET','Vérité UI','Registre visiteurs : affichait « CPR », enregistrait « CIN »', QA_REF),
 ('REJET','Vérité UI','Calendrier : en-têtes Lun–Dim sur une grille commençant dimanche', QA_REF),
 ('REJET','Vérité UI','« Joignable : Oui » codé en dur, sans aucune sonde', QA_REF),
 ('REJET','Argent','/app/finance CRASHAIT pour toute l\'école dès la 1re inscription réelle', QA_REF),
 ('REJET','Argent','Annuler une facture PAYÉE faisait disparaître l\'encaissé des états', QA_REF),
 ('REJET','Argent','money() : format français en anglais (12,345 lu « douze mille ») + décimales mixtes', QA_REF),
 ('REJET','Argent','Arrondi au dinar entier : chaque remise % volait ~234 fils', QA_REF),
 ('REJET','Légal','Inscription : « loi 2004-63 » (tunisienne) sous une case PDPL bahreïnie', QA_REF),
 ('REJET','Légal','Certificats et documents signés « (INPDP) » — autorité tunisienne', QA_REF),
 ('REJET','Données','Groupe sanguin « O+ » présélectionné ET ENREGISTRÉ (donnée médicale inventée)', QA_REF),
 ('REJET','Fonctionnel','Une CRÈCHE ne pouvait inscrire personne (niveaux 1ère–6ème seulement)', QA_REF),
 ('REJET','Fonctionnel','Formulaire élève : nom de 500 car., naissance 1900, tél « abc### » acceptés', QA_REF),
 ('REJET','RTL','Arabe : téléphones affichés À L\'ENVERS (« 111 111 20 216+ »), chiffres brouillés', QA_REF),
 ('REJET','Robustesse','Hors-ligne : chunk paresseux en échec = ÉCRAN BLANC muet', QA_REF),
 ('REJET','Démo','8 élèves homonymes par classe (« Chaima Karoui » ×8) — graine non crédible', QA_REF),
 ('MAJEUR','Fonctionnel','Appel : ré-enregistrer re-notifiait TOUS les parents', QA_REF),
 ('MAJEUR','Parent','Mes paiements : AUCUN montant, aucune devise', QA_REF),
 ('MAJEUR','Confirmations','Accident→parent, journal→parent, vider un jour de cantine : sans confirmation', QA_REF),
 ('MAJEUR','Rôles','RH et Comptabilité privés de Demandes / Espaces / Pointage (CR-019 incomplet)', QA_REF),
 ('MAJEUR','Rôles','RH ne pouvait pas décider un congé ; sécurité/RH/compta absents de l\'appel', QA_REF),
 ('MAJEUR','Autorisation','requests.js : assign/close SANS contrôle de rôle (garde uniquement à l\'écran)', QA_REF),
 ('MAJEUR','Honnêteté','Badge « EN DIRECT » sur une journée SIMULÉE montré au parent', QA_REF),
 ('MAJEUR','Admissions','Refuser un enfant : 1 clic, sans confirmation ni motif (le recrutement, lui, exige un motif)', QA_REF),
 ('MAJEUR','i18n','Toasts français derrière des boutons anglais (appel, moment, validation, envoi)', QA_REF),
 ('OUVERT','Argent','Deux modèles d\'argent (factures vs grille mensuelle) peuvent se contredire','OUVERT — D5'),
 ('OUVERT','Argent','Location et activités : encaissement sans reçu numéroté, hors comptabilité','OUVERT — D5'),
 ('OUVERT','Sécurité','Journal d\'audit inexistant (qui a lu un dossier médical) — exigence PDPL','OUVERT — D6'),
 ('OUVERT','i18n','~310 clés encore non traduites (EN 62 %, AR 63 %)','OUVERT — D4'),
 ('OUVERT','Exploitation','Hébergement serveur inexistant : données dans le navigateur, mots de passe en clair','OUVERT — D9'),
 ('OUVERT','Exploitation','Aucune supervision (Sentry/sonde) ni restauration de sauvegarde répétée','OUVERT — D9'),
 ('OUVERT','Pays','TVA 10 % / facture NBR · GOSI-EOSB dans la paie · reporting MOE','OUVERT — D3'),
 ('OUVERT','UX','Tuiles du tableau de bord n\'ouvrent pas toutes le détail ; pas d\'édition/archivage élève','OUVERT — D8'),
]
ws = sheet(wb, 'FAT 2026-07-26', 'FAT EXPLORATOIRE — 5 rapports QA (méthode « rejeter la livraison »)',
           'Chaque ingénieur QA a piloté un vrai navigateur, école configurée BAHREÏN, en EN et AR. Les 5 ont voté REJET.',
           ['Sévérité','Domaine','Constat','Statut'], [12,16,86,34])
fill(ws, FAT, sev_col=1)

# ═══ 12. NORME D'ACCEPTATION v1 ══════════════════════════════════════════════
ACC = [
 ('D1','Intégrité fonctionnelle','Aucun écran ne casse ; actions irréversibles confirmées ; aucune donnée inventée','core: npm test · e2e: npm run all · node e2e/fat.mjs','CONFORME','—'),
 ('D2','Vérité de l\'interface','Ce que l\'écran affirme est VRAI (bouton, identifiant, état mesuré)','Revue FAT exploratoire par rôle','CONFORME','—'),
 ('D3','Conformité pays','Semaine, devise, pièce, loi, urgences, fériés : du pack, jamais en dur','node e2e/diag.feries-pays.mjs · node e2e/fat.mjs','PARTIEL','TVA/NBR · GOSI-EOSB · reporting MOE'),
 ('D4','Langue FR/EN/AR','Aucune phrase d\'une autre langue à l\'écran','python3 docs/quality/i18n-audit.py','PARTIEL','EN 62 % / AR 63 % → seuil client 95 %'),
 ('D5','Intégrité de l\'argent','Rien ne disparaît ; arrondi au fils ; format par langue ; écritures confirmées','Tests cœur compta/paie + revue FAT argent','PARTIEL','Unifier les 2 modèles · reçus location/activités'),
 ('D6','Sécurité & données','Pas de secret, pas d\'élévation, accès par rôle, consentement pays, audit','securite.yml (CodeQL+secrets) · server tests','PARTIEL','Journal d\'audit · sortir du mode démo'),
 ('D7','Données du client','Entrée sans ressaisie, sortie libre','node e2e/diag.import.mjs · export OneRoster','CONFORME','(ZIP d\'export : confort)'),
 ('D8','UX & accessibilité','Pas d\'écran blanc, états complets, RTL, clavier, contraste','e2e parcours.qualite (axe) · parcours.uiux (25 pages)','CONFORME','—'),
 ('D9','Exploitation','3 environnements, tests bloquants, retour arrière, sauvegardes, supervision','deploy.yml + envs.yml','PARTIEL','HÉBERGEMENT SERVEUR · supervision · restauration testée'),
 ('D10','Preuve & documentation','Tout ce qui est affirmé est mesurable et daté','Ce classeur + ACCEPTANCE-STANDARD-v1.md','CONFORME','—'),
 ('','PRONONCÉ v1.5.0','PILOTE ACCOMPAGNÉ : oui — LIVRAISON AUTONOME : pas encore','','PILOTE','3 portes : D9 hébergement → D4 langue 95 % → D5/D3 argent & conformité'),
]
ws = sheet(wb, 'Norme d\'acceptation v1', 'NORME D\'ACCEPTATION PRODUIT v1 (2026-07-26)',
           'Ce que « acceptable pour une école qui paie » veut dire — critère, mesure exécutable, état. Détail : docs/quality/ACCEPTANCE-STANDARD-v1.md',
           ['#','Domaine','Critère','Comment on le mesure','État','Ce qui manque'], [6,26,52,44,12,50])
fill(ws, ACC, sev_col=5)

wb.save(OUT)
print(f'OK → {OUT}')
for s in wb.sheetnames:
    print(' ·', s)
